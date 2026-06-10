#!/usr/bin/env python3
"""
audit-log-writer.py
Writes one audit entry to the centralized skill audit log Excel file.
Called by skill-audit-logger at the end of every skill execution.

Usage:
  python audit-log-writer.py --init --path ~/Claude-Workspace/audit/skill-audit-log.xlsx
  python audit-log-writer.py --append --path ~/Claude-Workspace/audit/skill-audit-log.xlsx \
      --skill "adoption-playbook" \
      --proyecto "Nahual LSS" \
      --trigger "genera la guia de adopcion" \
      --idioma "español" \
      --entregables "adoption_guide.pptx" \
      --duracion 4 \
      --usuario "Sergio Ampudia" \
      --notas ""
"""

import argparse
import os
import sys
import json
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl --break-system-packages")
    sys.exit(1)


# ─── CONSTANTS ────────────────────────────────────────────────────────────────

HEADERS = [
    "fecha_hora", "skill_name", "proyecto", "trigger_phrase",
    "idioma", "entregables", "duracion_min", "usuario", "notas"
]

# Silanes branding
NAVY   = "1B3A6B"
RED    = "CC2222"
WHITE  = "FFFFFF"
LGRAY  = "E2E8F0"
OFFWH  = "F4F7FB"


# ─── STYLING ──────────────────────────────────────────────────────────────────

def style_header_row(ws):
    """Apply Silanes-branded header styling to row 1."""
    header_fill = PatternFill("solid", fgColor=NAVY)
    header_font = Font(bold=True, color=WHITE, name="Calibri", size=10)
    thin = Side(style="thin", color=LGRAY)
    border = Border(bottom=Side(style="medium", color=RED))

    for col_num, _ in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = border

    ws.row_dimensions[1].height = 22


def style_data_row(ws, row_num):
    """Alternate row shading."""
    fill_color = OFFWH if row_num % 2 == 0 else WHITE
    fill = PatternFill("solid", fgColor=fill_color)
    font = Font(name="Calibri", size=9)
    thin = Side(style="thin", color=LGRAY)
    border = Border(
        bottom=Side(style="thin", color=LGRAY),
        right=Side(style="thin", color=LGRAY)
    )
    for col_num in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = border

    ws.row_dimensions[row_num].height = 18


def set_column_widths(ws):
    widths = {
        "fecha_hora":     18,
        "skill_name":     22,
        "proyecto":       20,
        "trigger_phrase": 35,
        "idioma":         10,
        "entregables":    35,
        "duracion_min":   12,
        "usuario":        18,
        "notas":          40,
    }
    for col_num, header in enumerate(HEADERS, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = widths.get(header, 15)


# ─── RESUMEN SHEET ────────────────────────────────────────────────────────────

def update_resumen(wb):
    """Rebuild the Resumen pivot sheet from the Log data."""
    log_ws = wb["Log"]

    # Remove and recreate
    if "Resumen" in wb.sheetnames:
        del wb["Resumen"]
    rs = wb.create_sheet("Resumen")

    # Read all log data (skip header row 1)
    rows = list(log_ws.iter_rows(min_row=2, values_only=True))
    rows = [r for r in rows if r[0] is not None]  # skip blank rows

    if not rows:
        rs.cell(1, 1).value = "Sin datos aún."
        return

    # Index columns
    idx = {h: i for i, h in enumerate(HEADERS)}

    # ── By skill_name ──
    from collections import defaultdict
    skill_counts = defaultdict(int)
    skill_last   = {}
    for row in rows:
        sn = row[idx["skill_name"]] or "desconocido"
        skill_counts[sn] += 1
        ts = row[idx["fecha_hora"]]
        if sn not in skill_last or (ts and ts > skill_last[sn]):
            skill_last[sn] = ts

    # ── By project ──
    proj_counts = defaultdict(int)
    for row in rows:
        p = row[idx["proyecto"]] or "sin proyecto"
        proj_counts[p] += 1

    # ── By language ──
    lang_counts = defaultdict(int)
    for row in rows:
        l = row[idx["idioma"]] or "no especificado"
        lang_counts[l] += 1

    # ── This month ──
    now = datetime.now()
    this_month = sum(
        1 for r in rows
        if r[idx["fecha_hora"]] and
        str(r[idx["fecha_hora"]])[:7] == now.strftime("%Y-%m")
    )

    # ── Write Resumen ──
    header_fill = PatternFill("solid", fgColor=NAVY)
    header_font = Font(bold=True, color=WHITE, name="Calibri", size=10)
    section_font = Font(bold=True, color=NAVY, name="Calibri", size=10)

    def write_section(start_row, title, col_headers, data_rows):
        # Section title
        cell = rs.cell(start_row, 1, title)
        cell.font = section_font
        rs.row_dimensions[start_row].height = 20
        start_row += 1

        # Column headers
        for c, h in enumerate(col_headers, 1):
            cell = rs.cell(start_row, c, h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="left")
        rs.row_dimensions[start_row].height = 18
        start_row += 1

        # Data
        for dr in data_rows:
            for c, v in enumerate(dr, 1):
                cell = rs.cell(start_row, c, v)
                cell.font = Font(name="Calibri", size=9)
                fill_color = OFFWH if start_row % 2 == 0 else WHITE
                cell.fill = PatternFill("solid", fgColor=fill_color)
            rs.row_dimensions[start_row].height = 16
            start_row += 1

        return start_row + 1  # blank row between sections

    row = 1

    # Total summary box
    rs.cell(row, 1, "RESUMEN DE USO — SKILL AUDIT LOG").font = Font(
        bold=True, color=WHITE, name="Calibri", size=12
    )
    rs.cell(row, 1).fill = PatternFill("solid", fgColor=RED)
    rs.row_dimensions[row].height = 22
    row += 1

    rs.cell(row, 1, f"Total ejecuciones: {len(rows)}").font = Font(name="Calibri", size=10)
    rs.cell(row, 2, f"Este mes: {this_month}").font = Font(name="Calibri", size=10)
    rs.cell(row, 3, f"Generado: {now.strftime('%Y-%m-%d %H:%M')}").font = Font(
        name="Calibri", size=9, color="888888", italic=True
    )
    row += 2

    # By skill
    skill_data = sorted(
        [(sn, cnt, str(skill_last.get(sn, ""))[:16])
         for sn, cnt in skill_counts.items()],
        key=lambda x: -x[1]
    )
    row = write_section(row, "Ejecuciones por Skill",
                        ["skill_name", "total_runs", "ultima_ejecucion"],
                        skill_data)

    # By project
    proj_data = sorted([(p, c) for p, c in proj_counts.items()], key=lambda x: -x[1])
    row = write_section(row, "Ejecuciones por Proyecto",
                        ["proyecto", "total_runs"], proj_data)

    # By language
    lang_data = sorted([(l, c) for l, c in lang_counts.items()], key=lambda x: -x[1])
    row = write_section(row, "Ejecuciones por Idioma",
                        ["idioma", "total_runs"], lang_data)

    # Column widths for Resumen
    rs.column_dimensions["A"].width = 30
    rs.column_dimensions["B"].width = 14
    rs.column_dimensions["C"].width = 20


# ─── CORE FUNCTIONS ───────────────────────────────────────────────────────────

def init_log(path: Path):
    """Create a new audit log Excel file with headers and empty Log sheet."""
    path.parent.mkdir(parents=True, exist_ok=True)

    if path.exists():
        print(f"INFO: File already exists at {path}. No changes made.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Log"

    # Write headers
    for col_num, header in enumerate(HEADERS, 1):
        ws.cell(row=1, column=col_num, value=header)

    style_header_row(ws)
    set_column_widths(ws)
    ws.freeze_panes = "A2"

    # Create empty Resumen sheet
    rs = wb.create_sheet("Resumen")
    rs.cell(1, 1, "Sin datos aún. Ejecuta un skill para generar el resumen.")

    wb.save(path)
    print(f"✅ Audit log created: {path}")


def append_entry(path: Path, entry: dict):
    """Append one row to the Log sheet and update Resumen."""
    if not path.exists():
        print(f"WARNING: Log file not found at {path}. Creating it now.")
        init_log(path)

    wb = load_workbook(path)

    if "Log" not in wb.sheetnames:
        wb.create_sheet("Log", 0)
        ws = wb["Log"]
        for col_num, header in enumerate(HEADERS, 1):
            ws.cell(row=1, column=col_num, value=header)
        style_header_row(ws)
        set_column_widths(ws)
        ws.freeze_panes = "A2"
    else:
        ws = wb["Log"]

    # Find next empty row
    next_row = ws.max_row + 1
    if next_row == 2 and ws.cell(2, 1).value is None:
        next_row = 2  # file was just created

    # Write entry
    for col_num, header in enumerate(HEADERS, 1):
        ws.cell(row=next_row, column=col_num, value=entry.get(header, ""))

    style_data_row(ws, next_row)

    # Update Resumen
    update_resumen(wb)

    wb.save(path)
    print(f"✅ Entry logged (row {next_row}): {entry.get('skill_name')} | {entry.get('fecha_hora')}")


def save_fallback(path: Path, entry: dict):
    """Save entry as JSON sidecar if Excel write fails."""
    fallback = path.parent / f"audit-fallback-{datetime.now().strftime('%Y%m%d-%H%M%S')}.json"
    with open(fallback, "w", encoding="utf-8") as f:
        json.dump(entry, f, ensure_ascii=False, indent=2)
    print(f"⚠️  Excel write failed. Entry saved to fallback: {fallback}")


# ─── CLI ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Skill Audit Log Writer")
    parser.add_argument("--init",   action="store_true", help="Create new log file")
    parser.add_argument("--append", action="store_true", help="Append one entry")
    parser.add_argument("--path",   required=True,       help="Path to audit log Excel file")

    # Entry fields
    parser.add_argument("--skill",       default="", help="skill_name")
    parser.add_argument("--proyecto",    default="", help="proyecto")
    parser.add_argument("--trigger",     default="", help="trigger_phrase")
    parser.add_argument("--idioma",      default="", help="idioma")
    parser.add_argument("--entregables", default="", help="entregables (comma-separated)")
    parser.add_argument("--duracion",    default="", help="duracion_min (number)")
    parser.add_argument("--usuario",     default="", help="usuario")
    parser.add_argument("--notas",       default="", help="notas")

    args = parser.parse_args()
    path = Path(os.path.expanduser(args.path))

    if args.init:
        init_log(path)
        return

    if args.append:
        entry = {
            "fecha_hora":     datetime.now().strftime("%Y-%m-%d %H:%M"),
            "skill_name":     args.skill,
            "proyecto":       args.proyecto,
            "trigger_phrase": args.trigger,
            "idioma":         args.idioma,
            "entregables":    args.entregables,
            "duracion_min":   int(args.duracion) if args.duracion.isdigit() else args.duracion,
            "usuario":        args.usuario,
            "notas":          args.notas,
        }
        try:
            append_entry(path, entry)
        except Exception as e:
            print(f"ERROR writing to Excel: {e}")
            save_fallback(path, entry)
        return

    parser.print_help()


if __name__ == "__main__":
    main()
